from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path("/Users/nicholaslabriola/Documents/New project")
SOURCE_URL = "https://vote.sos.ri.gov/Forms/elections/Reports/Candidates.xlsx"
SOURCE = ROOT / "data" / "raw" / "ri_declared_candidates_official_2026.xlsx"
OUTPUT = ROOT / "website" / "data" / "whos_running_2026.json"

PRIMARY_ELECTION = "09/09/2026 - STATEWIDE PRIMARY"
GENERAL_ELECTION = "11/03/2026 - STATEWIDE GENERAL ELECTION"
LEGISLATIVE_OFFICES = {
    "REPRESENTATIVE IN GENERAL ASSEMBLY": "house",
    "SENATOR IN GENERAL ASSEMBLY": "senate",
}

CHAMBER_TOTALS = {"house": 75, "senate": 38}

PARTY_CODE = {
    "democrat": "DEM",
    "democratic": "DEM",
    "republican": "REP",
    "independent": "IND",
    "independent-socialist": "OTH",
    "libertarian": "LIB",
    "moderate": "MOD",
}

PARTY_LABEL = {
    "DEM": "Democratic",
    "REP": "Republican",
    "IND": "Independent",
    "MOD": "Moderate",
    "LIB": "Libertarian",
    "OTH": "Other",
}

ALLOWED_DECLARATION_STATUSES = {"Valid"}


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_zip(value: Any) -> str:
    clean = normalize_space(value)
    digits = "".join(ch for ch in clean if ch.isdigit())
    return digits[:5] if digits else ""


def normalize_city(value: Any) -> str:
    clean = normalize_space(value).lower()
    if not clean:
        return ""
    replacements = {
        "n ": "north ",
        "s ": "south ",
        "e ": "east ",
        "w ": "west ",
        "mt ": "mount ",
        "st ": "saint ",
    }
    for short, full in replacements.items():
        if clean.startswith(short):
            clean = full + clean[len(short):]
    return " ".join(word.capitalize() for word in clean.split())


def normalize_party(value: Any) -> str:
    clean = normalize_space(value).lower()
    return PARTY_CODE.get(clean, "OTH")


def clean_name_part(value: Any) -> str:
    return normalize_space(value)


def build_display_name(first: Any, middle: Any, last: Any, suffix: Any) -> str:
    parts = [
        clean_name_part(first),
        clean_name_part(middle),
        clean_name_part(last),
        clean_name_part(suffix),
    ]
    return " ".join(part for part in parts if part)


def bool_from_yes_no(value: Any) -> bool | None:
    clean = normalize_space(value).lower()
    if clean == "yes":
        return True
    if clean == "no":
        return False
    return None


def parse_int(value: Any) -> int | None:
    clean = normalize_space(value).replace(",", "")
    if not clean:
        return None
    try:
        return int(clean)
    except ValueError:
        return None


def download_source_workbook() -> dict[str, str]:
    SOURCE.parent.mkdir(parents=True, exist_ok=True)
    request = Request(
        SOURCE_URL,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )

    downloaded_at = datetime.now(timezone.utc).isoformat()
    with urlopen(request, timeout=45) as response:
        SOURCE.write_bytes(response.read())
        last_modified = response.headers.get("Last-Modified")

    source_updated_at = ""
    if last_modified:
        source_updated_at = parsedate_to_datetime(last_modified).astimezone(timezone.utc).isoformat()
    else:
        source_updated_at = datetime.fromtimestamp(SOURCE.stat().st_mtime, timezone.utc).isoformat()

    return {
        "source_downloaded_at": downloaded_at,
        "source_updated_at": source_updated_at,
    }


def build_candidate(row: dict[str, Any], ballot_stage: str) -> dict[str, Any]:
    party_raw = normalize_space(row["PARTY"])
    party = normalize_party(party_raw)
    city = normalize_city(row["POSTAL CITY"])
    zip_code = normalize_zip(row["ZIP CODE"])
    home_label = city if not zip_code else f"{city} {zip_code}"

    return {
        "name": build_display_name(row["FIRST NAME"], row["MIDDLE NAME"], row["LAST NAME"], row["SUFFIX"]),
        "first_name": clean_name_part(row["FIRST NAME"]),
        "middle_name": clean_name_part(row["MIDDLE NAME"]),
        "last_name": clean_name_part(row["LAST NAME"]),
        "suffix": clean_name_part(row["SUFFIX"]),
        "party": party,
        "party_label": PARTY_LABEL.get(party, "Other"),
        "party_raw": party_raw,
        "hometown": city,
        "zip_code": zip_code,
        "home_label": home_label,
        "phone": normalize_space(row["PHONE#"]),
        "email": normalize_space(row["EMAIL"]).lower(),
        "declaration_status": normalize_space(row["DECLARATION"]),
        "needs_nomination_papers": bool_from_yes_no(row["NEED N.P."]),
        "required_signatures": parse_int(row["REQ"]),
        "official_source": "Rhode Island Department of State",
        "ballot_stage": ballot_stage,
    }


def upsert_candidate(candidates: list[dict[str, Any]], candidate: dict[str, Any]) -> None:
    existing = next(
        (
            item
            for item in candidates
            if item["name"] == candidate["name"]
            and item["party"] == candidate["party"]
            and item["ballot_stage"] == candidate["ballot_stage"]
        ),
        None,
    )
    if existing:
        for key, value in candidate.items():
            if value not in ("", None):
                existing[key] = value
        return
    candidates.append(candidate)


def build_status(counts: dict[str, int]) -> tuple[str, str]:
    dem = counts.get("DEM", 0)
    rep = counts.get("REP", 0)
    other = sum(v for k, v in counts.items() if k not in {"DEM", "REP"})

    if dem and rep:
        general = "both_major"
    elif dem:
        general = "dem_only"
    elif rep:
        general = "rep_only"
    elif other:
        general = "other_only"
    else:
        general = "no_filing"

    dem_primary = dem > 1
    rep_primary = rep > 1

    if dem_primary and rep_primary:
        primary = "both_primaries"
    elif dem_primary:
        primary = "dem_primary"
    elif rep_primary:
        primary = "rep_primary"
    elif sum(counts.values()) > 0:
        primary = "no_primary"
    else:
        primary = "no_filing"

    return general, primary


def general_label(code: str) -> str:
    return {
        "both_major": "Democratic and Republican candidates filed",
        "dem_only": "Democratic candidates filed only",
        "rep_only": "Republican candidates filed only",
        "other_only": "Independent or other candidates filed only",
        "no_filing": "No primary filing listed in official state workbook",
    }[code]


def primary_label(code: str) -> str:
    return {
        "both_primaries": "Both parties have primary activity",
        "dem_primary": "Democratic primary underway",
        "rep_primary": "Republican primary underway",
        "no_primary": "No active primary in current filings",
        "no_filing": "No primary filing listed in official state workbook",
    }[code]


def main() -> None:
    source_meta = download_source_workbook()
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook["CandidateElection"]
    headers = [cell.value for cell in sheet[1]]
    rows = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

    chamber_records: dict[str, dict[int, dict[str, Any]]] = {
        "house": defaultdict(lambda: {"candidates": [], "general_candidates": [], "party_counts": defaultdict(int)}),
        "senate": defaultdict(lambda: {"candidates": [], "general_candidates": [], "party_counts": defaultdict(int)}),
    }

    for row in rows:
        office = normalize_space(row.get("OFFICE"))
        chamber = LEGISLATIVE_OFFICES.get(office)
        if not chamber:
            continue

        election_name = normalize_space(row.get("ELECTION DATE - NAME"))
        declaration_status = normalize_space(row.get("DECLARATION"))
        if declaration_status not in ALLOWED_DECLARATION_STATUSES:
            continue

        district = parse_int(row.get("DIST#"))
        if district is None:
            continue

        record = chamber_records[chamber][district]

        if election_name == PRIMARY_ELECTION:
            candidate = build_candidate(row, ballot_stage="primary")
            upsert_candidate(record["candidates"], candidate)
            record["party_counts"][candidate["party"]] += 1
        elif election_name == GENERAL_ELECTION:
            candidate = build_candidate(row, ballot_stage="general")
            upsert_candidate(record["general_candidates"], candidate)

    payload: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "election_cycle": 2026,
        "source_kind": "official_ri_declared_candidates_workbook",
        "source_download_url": SOURCE_URL,
        "source_file": str(SOURCE),
        "source_downloaded_at": source_meta["source_downloaded_at"],
        "source_updated_at": source_meta["source_updated_at"],
        "manual_cross_check_notes": [
            "Primary map colors and district counts are based on the official Rhode Island Department of State declared candidates workbook.",
            "Only legislative candidates with DECLARATION = Valid are included in the published dataset.",
            "Street addresses from the official workbook are intentionally withheld on the public site; hometown, ZIP code, phone, and email are shown when available.",
            "General-election-only independents are preserved separately in district detail cards but do not color the September 9 primary map.",
        ],
        "chambers": {},
        "summary": {},
    }

    for chamber, districts in chamber_records.items():
        chamber_payload = {}
        filed_districts = 0
        both_major = 0
        primary_districts = 0
        dem_only = 0
        rep_only = 0
        general_only_districts = 0
        empty = CHAMBER_TOTALS[chamber]

        for district in range(1, CHAMBER_TOTALS[chamber] + 1):
            raw = districts.get(district)
            counts = dict(raw["party_counts"]) if raw else {}
            primary_candidates = sorted(
                raw["candidates"] if raw else [],
                key=lambda item: (item["party_label"], item["name"]),
            )
            general_candidates = sorted(
                raw["general_candidates"] if raw else [],
                key=lambda item: (item["party_label"], item["name"]),
            )
            general_status, primary_status = build_status(counts)
            total_candidates = sum(counts.values())

            if total_candidates:
                filed_districts += 1
                empty -= 1
            if general_status == "both_major":
                both_major += 1
            elif general_status == "dem_only":
                dem_only += 1
            elif general_status == "rep_only":
                rep_only += 1
            if primary_status in {"dem_primary", "rep_primary", "both_primaries"}:
                primary_districts += 1
            if general_candidates:
                general_only_districts += 1

            hometowns = sorted(
                {
                    candidate["home_label"]
                    for candidate in [*primary_candidates, *general_candidates]
                    if candidate.get("home_label")
                }
            )

            chamber_payload[str(district)] = {
                "chamber": chamber,
                "district_number": district,
                "candidates": primary_candidates,
                "general_candidates": general_candidates,
                "party_counts": counts,
                "candidate_total": total_candidates,
                "general_candidate_total": len(general_candidates),
                "general_status": general_status,
                "general_label": general_label(general_status),
                "primary_status": primary_status,
                "primary_label": primary_label(primary_status),
                "hometowns": hometowns,
            }

        payload["chambers"][chamber] = chamber_payload
        payload["summary"][chamber] = {
            "district_total": CHAMBER_TOTALS[chamber],
            "districts_with_filings": filed_districts,
            "no_filing_reported": empty,
            "both_major_parties_filed": both_major,
            "democratic_only": dem_only,
            "republican_only": rep_only,
            "districts_with_primary_activity": primary_districts,
            "districts_with_general_only_candidates": general_only_districts,
        }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
